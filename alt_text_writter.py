# import os
# import requests
# from bs4 import BeautifulSoup
# from openpyxl import Workbook
# from openpyxl.drawing.image import Image
# from PIL import Image as PILImage
# import csv

# # Input and output file names
# input_csv = r"C:\Users\SuryaRaghupathy\OneDrive - Nurtur Limited\Desktop\Git code import\alt_text_blogs_lists.csv"

# output_excel = "image_data.xlsx"  # Output Excel file

# # Base directory for saving images
# base_directory = r"C:\Users\SuryaRaghupathy\OneDrive - Nurtur Limited\Desktop\Git code import"

# # Headers to mimic a browser
# headers = {
#     "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36"
# }

# # Resize images for better clarity in sheets
# def resize_image(image_path, max_width, max_height):
#     with PILImage.open(image_path) as img:
#         img.thumbnail((max_width, max_height))
#         img.save(image_path)

# # Create a new workbook
# workbook = Workbook()
# sheet = workbook.active
# sheet.append(["Page URL", "Image Source URL", "Image File Path"])  # Add headers

# # Read URLs from the input CSV file
# with open(input_csv, mode='r', encoding='utf-8') as csv_file:
#     reader = csv.DictReader(csv_file)
    
#     for row in reader:
#         url = row["Page URL"]  # Assuming the column name is 'Page URL'
#         print(f"Processing URL: {url}")
        
#         try:
#             # Send a request to the webpage
#             response = requests.get(url, headers=headers)
#             response.raise_for_status()  # Ensure the request was successful
            
#             # Parse the page content
#             soup = BeautifulSoup(response.text, "html.parser")
            
#             # Find the img tag with a class name partially matching "attachment-large size-large wp-image"
#             img_tag = soup.find("img", class_=lambda x: x and "single-banner-image w-100 object-fit-cover img-fluid wp-post-image" in x)
            
#             # Initialize variables for this row
#             img_src = None
#             image_filename = None
            
#             # Get the src attribute of the img tag if it exists
#             if img_tag and "src" in img_tag.attrs:
#                 img_src = img_tag["src"]
#                 print(f"Image source URL: {img_src}")
                
#                 # Download the image
#                 img_response = requests.get(img_src, headers=headers, stream=True)
#                 img_response.raise_for_status()
                
#                 # Save the image locally
#                 image_name = img_src.split("/")[-1]
#                 image_path = os.path.join(base_directory, image_name)
#                 with open(image_path, "wb") as img_file:
#                     img_file.write(img_response.content)
                
#                 # Resize the image for clarity
#                 resize_image(image_path, 300, 300)
#                 print(f"Image downloaded, resized, and saved as {image_path}")
            
#             # Append data to the Excel file
#             sheet.append([url, img_src if img_src else "No image found", image_path if img_src else "No image found"])
            
#             # Embed the image in the Excel sheet if it was downloaded
#             if img_src:
#                 img = Image(image_path)
#                 img.anchor = f"C{sheet.max_row}"  # Adjust cell placement
#                 sheet.add_image(img)
        
#         except Exception as e:
#             print(f"Error processing URL {url}: {e}")
#             sheet.append([url, "Error", ""])

# # Save the workbook
# workbook.save(output_excel)
# print(f"Data and images have been saved to {output_excel}")
import os
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import csv

# Input and output file names
input_csv = r"C:\Users\SuryaRaghupathy\OneDrive - Nurtur Limited\Desktop\Git code import\alt_text_blogs_lists.csv"
output_excel = "image_data.xlsx"  # Output Excel file

# Base directory for saving images
base_directory = r"C:\Users\SuryaRaghupathy\OneDrive - Nurtur Limited\Desktop\Git code import"

# Headers to mimic a browser
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36"
}

# Resize images for better clarity in sheets
def resize_image(image_path, max_width, max_height):
    try:
        with PILImage.open(image_path) as img:
            img.thumbnail((max_width, max_height))
            img.save(image_path)
    except Exception as e:
        print(f"Error resizing image {image_path}: {e}")

# Create a new workbook
workbook = Workbook()
sheet = workbook.active
sheet.append(["Page URL", "Image Source URL", "Image File Path"])  # Add headers

# Read URLs from the input CSV file
with open(input_csv, mode='r', encoding='utf-8') as csv_file:
    reader = csv.DictReader(csv_file)

    for row in reader:
        url = row["Page URL"]
        print(f"\nProcessing URL: {url}")

        try:
            # Send a request to the webpage
            response = requests.get(url, headers=headers)
            response.raise_for_status()

            # Parse the page content
            soup = BeautifulSoup(response.text, "html.parser")

            # Find all <img> tags
            img_tags = soup.find_all("img")

            if not img_tags:
                print(f"No images found at {url}")
                sheet.append([url, "No images found", ""])
                continue

            for img_tag in img_tags:
                img_src = img_tag.get("src")

                if not img_src or not img_src.startswith("http"):
                    continue  # Skip if no src or if it's a relative path

                try:
                    print(f"Downloading image: {img_src}")
                    img_response = requests.get(img_src, headers=headers, stream=True)
                    img_response.raise_for_status()

                    # Save the image
                    image_name = img_src.split("/")[-1].split("?")[0]
                    image_path = os.path.join(base_directory, image_name)

                    with open(image_path, "wb") as img_file:
                        img_file.write(img_response.content)

                    # Resize the image
                    resize_image(image_path, 300, 300)

                    # Write to Excel
                    sheet.append([url, img_src, image_path])

                    # Embed image in Excel
                    img = Image(image_path)
                    img.anchor = f"C{sheet.max_row}"
                    sheet.add_image(img)

                except Exception as img_error:
                    print(f"Error downloading/processing image {img_src}: {img_error}")
                    sheet.append([url, img_src, "Download error"])

        except Exception as e:
            print(f"Error processing URL {url}: {e}")
            sheet.append([url, "Page error", ""])

# Save the workbook
workbook.save(output_excel)
print(f"\n✅ Data and images have been saved to {output_excel}")
